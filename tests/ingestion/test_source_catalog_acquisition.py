from __future__ import annotations

import csv
import json
import zipfile
from typing import TYPE_CHECKING

from openpyxl import Workbook

from bubble.analysis.source_coverage import build_source_coverage_report
from bubble.ingestion.sources import SourceCatalogClient, acquire_source_catalog

if TYPE_CHECKING:
    from pathlib import Path


def _write_csv(path: Path, rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    assert rows
    fieldnames = list(dict.fromkeys(key for row in rows for key in row))
    with path.open("w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def test_source_catalog_acquires_raw_artifacts_and_normalized_rows(tmp_path: Path):
    queue_source = tmp_path / "ercot_queue.csv"
    permit_source = tmp_path / "permits.json"
    ppa_source = tmp_path / "ppa.txt"
    _write_csv(
        queue_source,
        [{"queue_id": "Q-1", "project_id": "campus-1", "mw": 250, "status": "active"}],
    )
    permit_source.write_text(
        json.dumps({"records": [{"permit_id": "AIR-1", "project_id": "campus-1"}]})
    )
    ppa_source.write_text("Power Purchase Agreement for Campus 1, 200 MW.")
    catalog = tmp_path / "source_catalog.csv"
    _write_csv(
        catalog,
        [
            {
                "source_id": "ercot-q",
                "corpus": "queue_records",
                "source_uri": queue_source.as_uri(),
                "source_type": "grid_interconnection_queue",
                "parser": "csv",
                "project_id": "campus-1",
                "meta_region": "ERCOT",
            },
            {
                "source_id": "air-permit",
                "corpus": "permit_records",
                "source_uri": permit_source.as_uri(),
                "source_type": "state_deq",
                "parser": "json",
                "project_id": "campus-1",
            },
            {
                "source_id": "ppa-doc",
                "corpus": "ppas",
                "source_uri": ppa_source.as_uri(),
                "source_type": "state_puc",
                "parser": "text",
                "project_id": "campus-1",
            },
        ],
    )

    batch = acquire_source_catalog(
        catalog,
        output_dir=tmp_path / "acquired",
        max_workers=12,
        sec_requests_per_second=7.5,
        sec_domain_concurrency=5,
        other_requests_per_second=24.0,
        other_domain_concurrency=9,
        retry_attempts=5,
    )

    assert batch.summary.catalog_rows == 3
    assert batch.summary.artifacts_acquired == 3
    assert batch.summary.extracted_rows == 3
    assert batch.summary.workers == 3
    assert batch.summary.sec_requests_per_second == 7.5
    assert batch.summary.sec_domain_concurrency == 5
    assert batch.summary.other_requests_per_second == 24.0
    assert batch.summary.other_domain_concurrency == 9
    assert batch.summary.retry_attempts == 5
    assert batch.summary.resume_enabled is True
    assert batch.summary.corpora == {"permit_records": 1, "ppas": 1, "queue_records": 1}

    inventory = tmp_path / "acquired" / "source_artifact_inventory.csv"
    queue_rows = tmp_path / "acquired" / "source_rows" / "queue_records.csv"
    permit_rows = tmp_path / "acquired" / "source_rows" / "permit_records.csv"
    ppa_rows = tmp_path / "acquired" / "source_rows" / "ppas.csv"
    assert inventory.exists()
    assert queue_rows.exists()
    assert permit_rows.exists()
    assert ppa_rows.exists()
    assert "content_hash" in queue_rows.read_text().splitlines()[0]
    assert "retrieved_at" in permit_rows.read_text().splitlines()[0]
    assert "Power Purchase Agreement" in ppa_rows.read_text()

    coverage = build_source_coverage_report([tmp_path / "acquired"])
    assert coverage.source_documents == 3
    assert coverage.projects == 1
    assert coverage.queue_records == 1
    assert coverage.permit_records == 1
    assert coverage.ppas == 1
    assert coverage.source_backed_deals == 1
    assert coverage.deal_types == {"ppa": 1}


def test_source_catalog_client_requires_identity_for_sec_sources(monkeypatch):
    monkeypatch.delenv("EDGAR_IDENTITY", raising=False)
    client = SourceCatalogClient(identity=None)

    try:
        client.fetch_bytes("https://data.sec.gov/submissions/CIK0000789019.json")
    except ValueError as exc:
        assert "EDGAR_IDENTITY" in str(exc)
    else:  # pragma: no cover - defensive assertion for accidental network access
        raise AssertionError("SEC acquisition should require explicit identity")


def test_source_catalog_json_records_path_selects_nested_records(tmp_path: Path):
    source = tmp_path / "page.json"
    source.write_text(
        json.dumps({"rowData": [{"ID": "1", "Counterparty_Name": "Utility"}], "totalCount": 1})
    )
    catalog = tmp_path / "source_catalog.csv"
    _write_csv(
        catalog,
        [
            {
                "source_id": "ferc-page",
                "corpus": "ppas",
                "source_uri": source.as_uri(),
                "source_type": "ferc",
                "parser": "json",
                "meta_json_records_path": "rowData",
            }
        ],
    )

    batch = acquire_source_catalog(catalog, output_dir=tmp_path / "acquired")

    assert batch.summary.extracted_rows == 1
    ppa_rows = (tmp_path / "acquired" / "source_rows" / "ppas.csv").read_text()
    assert "Counterparty_Name" in ppa_rows
    assert "Utility" in ppa_rows


def test_source_catalog_flattens_arcgis_feature_records(tmp_path: Path):
    source = tmp_path / "arcgis-page.json"
    source.write_text(
        json.dumps(
            {
                "features": [
                    {
                        "attributes": {
                            "facility_id": "FT-1",
                            "facility_name": "Data Campus",
                            "operator_name": "Operator LLC",
                        },
                        "geometry": {"x": -87.0, "y": 33.0},
                    }
                ]
            }
        )
    )
    catalog = tmp_path / "source_catalog.csv"
    _write_csv(
        catalog,
        [
            {
                "source_id": "fractracker-page",
                "corpus": "tracker_records",
                "source_uri": source.as_uri(),
                "source_type": "project_tracker",
                "parser": "json",
                "meta_json_records_path": "features",
                "meta_json_flatten_records": "true",
            }
        ],
    )

    batch = acquire_source_catalog(catalog, output_dir=tmp_path / "acquired")

    assert batch.summary.extracted_rows == 1
    tracker_rows = (tmp_path / "acquired" / "source_rows" / "tracker_records.csv").read_text()
    assert "facility_name" in tracker_rows
    assert "Data Campus" in tracker_rows
    assert "operator_name" in tracker_rows
    assert "geometry_x" in tracker_rows

    coverage = build_source_coverage_report([tmp_path / "acquired"])
    assert coverage.tracker_records == 1
    assert coverage.projects == 1


def test_source_catalog_parses_xlsx_rows(tmp_path: Path):
    workbook_path = tmp_path / "queue.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Queue"
    worksheet.append(["queue_id", "project_id", "mw"])
    worksheet.append(["Q-1", "campus-1", 250])
    worksheet.append(["Q-2", "campus-2", 100])
    workbook.save(workbook_path)

    catalog = tmp_path / "source_catalog.csv"
    _write_csv(
        catalog,
        [
            {
                "source_id": "queue-xlsx",
                "corpus": "queue_records",
                "source_uri": workbook_path.as_uri(),
                "source_type": "grid_interconnection_queue",
                "parser": "xlsx",
            }
        ],
    )

    batch = acquire_source_catalog(catalog, output_dir=tmp_path / "acquired")

    assert batch.summary.extracted_rows == 2
    queue_rows = tmp_path / "acquired" / "source_rows" / "queue_records.csv"
    text = queue_rows.read_text()
    assert "Q-1" in text
    assert "campus-2" in text
    assert "content_hash" in text.splitlines()[0]


def test_source_catalog_parses_csv_after_metadata_rows(tmp_path: Path):
    source = tmp_path / "spp.csv"
    source.write_text(
        '"Last Updated On",5/29/2026,\nGeneration Interconnection Number,Capacity\nTI-18-0827,145\n'
    )
    catalog = tmp_path / "source_catalog.csv"
    _write_csv(
        catalog,
        [
            {
                "source_id": "spp-active",
                "corpus": "queue_records",
                "source_uri": source.as_uri(),
                "source_type": "grid_interconnection_queue",
                "parser": "csv",
                "meta_csv_skip_rows": "1",
            }
        ],
    )

    batch = acquire_source_catalog(catalog, output_dir=tmp_path / "acquired")

    assert batch.summary.extracted_rows == 1
    queue_rows = tmp_path / "acquired" / "source_rows" / "queue_records.csv"
    text = queue_rows.read_text()
    assert "TI-18-0827" in text
    assert "Last Updated" not in text.splitlines()[0]


def test_source_catalog_parses_xml_rows(tmp_path: Path):
    source = tmp_path / "pjm.xml"
    source.write_text(
        """<?xml version="1.0" encoding="UTF-8"?>
        <Projects>
          <Project>
            <ProjectNumber>Q-1</ProjectNumber>
            <Name>Data Center Substation</Name>
            <MaximumFacilityOutput>450</MaximumFacilityOutput>
            <Status>Active</Status>
          </Project>
          <Project>
            <ProjectNumber>Q-2</ProjectNumber>
            <Name>Withdrawn Project</Name>
            <MaximumFacilityOutput>100</MaximumFacilityOutput>
            <Status>Withdrawn</Status>
          </Project>
        </Projects>
        """
    )
    catalog = tmp_path / "source_catalog.csv"
    _write_csv(
        catalog,
        [
            {
                "source_id": "pjm-xml",
                "corpus": "queue_records",
                "source_uri": source.as_uri(),
                "source_type": "grid_interconnection_queue",
                "parser": "xml",
            }
        ],
    )

    batch = acquire_source_catalog(catalog, output_dir=tmp_path / "acquired")

    assert batch.summary.extracted_rows == 2
    queue_rows = tmp_path / "acquired" / "source_rows" / "queue_records.csv"
    text = queue_rows.read_text()
    assert "Data Center Substation" in text
    assert "MaximumFacilityOutput" in text.splitlines()[0]
    assert "content_hash" in text.splitlines()[0]


def test_source_catalog_parses_selected_zip_csv_members(tmp_path: Path):
    source = tmp_path / "epa.zip"
    with zipfile.ZipFile(source, "w") as archive:
        archive.writestr(
            "ICIS-AIR_FACILITIES.csv",
            "PGM_SYS_ID,FACILITY_NAME,STATE\nAIR-1,Campus Turbine Plant,VA\n",
        )
        archive.writestr(
            "ICIS-AIR_PROGRAMS.csv",
            "PGM_SYS_ID,PROGRAM_CODE,PROGRAM_DESC\nAIR-1,CAATVP,Title V Permit\n",
        )
        archive.writestr("IGNORED.csv", "id\n1\n")
    catalog = tmp_path / "source_catalog.csv"
    _write_csv(
        catalog,
        [
            {
                "source_id": "epa-icis-air",
                "corpus": "permit_records",
                "source_uri": source.as_uri(),
                "source_type": "epa",
                "parser": "zip",
                "meta_zip_member_names": "ICIS-AIR_FACILITIES.csv|ICIS-AIR_PROGRAMS.csv",
            }
        ],
    )

    batch = acquire_source_catalog(catalog, output_dir=tmp_path / "acquired")

    assert batch.summary.extracted_rows == 2
    permit_rows = tmp_path / "acquired" / "source_rows" / "permit_records.csv"
    text = permit_rows.read_text()
    assert "Campus Turbine Plant" in text
    assert "Title V Permit" in text
    assert "IGNORED" not in text
    assert "zip_member" in text.splitlines()[0]


def test_source_catalog_streams_selected_zip_xml_members(tmp_path: Path):
    source = tmp_path / "gleif.zip"
    with zipfile.ZipFile(source, "w") as archive:
        archive.writestr(
            "20260601-gleif-concatenated-file-rr.xml",
            """<?xml version="1.0" encoding="UTF-8"?>
            <rr:RelationshipData xmlns:rr="http://www.gleif.org/data/schema/rr/2016">
              <rr:RelationshipRecords>
                <rr:RelationshipRecord>
                  <rr:Relationship>
                    <rr:StartNode>
                      <rr:NodeID>CHILDLEI123</rr:NodeID>
                      <rr:NodeIDType>LEI</rr:NodeIDType>
                    </rr:StartNode>
                    <rr:EndNode>
                      <rr:NodeID>PARENTLEI456</rr:NodeID>
                      <rr:NodeIDType>LEI</rr:NodeIDType>
                    </rr:EndNode>
                    <rr:RelationshipType>IS_DIRECTLY_CONSOLIDATED_BY</rr:RelationshipType>
                    <rr:RelationshipStatus>ACTIVE</rr:RelationshipStatus>
                  </rr:Relationship>
                </rr:RelationshipRecord>
              </rr:RelationshipRecords>
            </rr:RelationshipData>
            """,
        )
    catalog = tmp_path / "source_catalog.csv"
    _write_csv(
        catalog,
        [
            {
                "source_id": "gleif-rr",
                "corpus": "ownership_records",
                "source_uri": source.as_uri(),
                "source_type": "gleif",
                "parser": "zip",
                "meta_zip_member_name_prefix": "20260601",
                "meta_zip_xml_record_tag": "RelationshipRecord",
            }
        ],
    )

    batch = acquire_source_catalog(catalog, output_dir=tmp_path / "acquired")

    assert batch.summary.extracted_rows == 1
    ownership_rows = tmp_path / "acquired" / "source_rows" / "ownership_records.csv"
    text = ownership_rows.read_text()
    assert "CHILDLEI123" in text
    assert "PARENTLEI456" in text
    assert "Relationship_StartNode_NodeID" in text.splitlines()[0]


def test_source_catalog_uses_explicit_file_extension_for_extensionless_urls(tmp_path: Path):
    catalog = tmp_path / "source_catalog.csv"
    _write_csv(
        catalog,
        [
            {
                "source_id": "iso-ne-export",
                "corpus": "queue_records",
                "source_uri": "https://example.com/reports/exportpublicqueue?ReportDate=1",
                "source_type": "grid_interconnection_queue",
                "parser": "text",
                "meta_file_extension": "xlsx",
            }
        ],
    )

    batch = acquire_source_catalog(
        catalog,
        output_dir=tmp_path / "acquired",
        fetch_bytes=lambda _url: b"export bytes",
    )

    assert batch.artifacts[0].local_path.endswith("iso-ne-export.xlsx")


def test_source_catalog_filters_xlsx_sheets_and_scans_deep_headers(tmp_path: Path):
    workbook_path = tmp_path / "ercot.xlsx"
    workbook = Workbook()
    contents = workbook.active
    contents.title = "Contents"
    contents.append(["Table of Contents"])
    contents.append(["not a queue record"])
    worksheet = workbook.create_sheet("Project Details - Large Gen")
    for _ in range(30):
        worksheet.append([None, None, None])
    worksheet.append(["INR", "Project Name", "Capacity (MW)"])
    worksheet.append([None, None, "subheader"])
    worksheet.append(["26INR0001", "West Texas Campus", 500])
    workbook.save(workbook_path)

    catalog = tmp_path / "source_catalog.csv"
    _write_csv(
        catalog,
        [
            {
                "source_id": "ercot-xlsx",
                "corpus": "queue_records",
                "source_uri": workbook_path.as_uri(),
                "source_type": "grid_interconnection_queue",
                "parser": "xlsx",
                "meta_xlsx_sheet_name_prefix": "Project Details -",
                "meta_xlsx_required_value_columns": "INR|Project Name",
            }
        ],
    )

    batch = acquire_source_catalog(catalog, output_dir=tmp_path / "acquired")

    assert batch.summary.extracted_rows == 1
    queue_rows = tmp_path / "acquired" / "source_rows" / "queue_records.csv"
    text = queue_rows.read_text()
    assert "26INR0001" in text
    assert "West Texas Campus" in text
    assert "not a queue record" not in text
    assert "subheader" not in text


def test_source_catalog_resume_reuses_existing_artifact(tmp_path: Path):
    source = tmp_path / "queue.csv"
    _write_csv(source, [{"queue_id": "Q-1", "project_id": "campus-1"}])
    catalog = tmp_path / "source_catalog.csv"
    _write_csv(
        catalog,
        [
            {
                "source_id": "queue-csv",
                "corpus": "queue_records",
                "source_uri": "https://example.com/queue.csv",
                "source_type": "grid_interconnection_queue",
                "parser": "csv",
            }
        ],
    )
    calls = 0

    def fake_fetch(_url: str) -> bytes:
        nonlocal calls
        calls += 1
        return source.read_bytes()

    first = acquire_source_catalog(
        catalog, output_dir=tmp_path / "acquired", fetch_bytes=fake_fetch
    )
    second = acquire_source_catalog(
        catalog, output_dir=tmp_path / "acquired", fetch_bytes=fake_fetch
    )

    assert calls == 1
    assert first.summary.artifacts_resumed == 0
    assert second.summary.artifacts_resumed == 1
