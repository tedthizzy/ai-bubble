"""CAISO interconnection-queue ingestion (off-EDGAR, KEYLESS — corrects the premature 'wall').

Proves the ISO-queue dimension is reachable without a key/payment (I'd wrongly written it
off after one 403 on a different host). Downloads the CAISO public queue report and
quantifies the active / completed / WITHDRAWN generation pipeline — the withdrawn pile is
the realized 'announced-but-cancelled' buildout, the power-layer analog of the fiber-1999
overbuild thesis (case zero's physical reality).

Output: analysis/iso_queue_caiso.{json,md}
"""

from __future__ import annotations

import io
import json
import urllib.request
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
OUT_JSON = ROOT / "analysis" / "iso_queue_caiso.json"
OUT_MD = ROOT / "analysis" / "iso_queue_caiso.md"
URL = "https://www.caiso.com/PublishedDocuments/PublicQueueReport.xlsx"
UA = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/120 Safari/537.36"


def find_header(ws) -> tuple[int, dict[int, str]]:
    best_row, best, best_map = 1, 0, {}
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True), 1):
        cells = [(i, str(c).strip()) for i, c in enumerate(row) if c not in (None, "")]
        strcells = [(i, v) for i, v in cells if any(ch.isalpha() for ch in v)]
        if len(strcells) > best:
            best, best_row, best_map = len(strcells), ri, {i: v for i, v in strcells}
    return best_row, best_map


def col_for(header: dict[int, str], *keys: str) -> int | None:
    for i, name in header.items():
        low = name.lower()
        if any(k in low for k in keys):
            return i
    return None


def summarize_sheet(ws) -> dict:
    hr, header = find_header(ws)
    mw_col = col_for(header, "mw", "capacity", "net mws")
    type_col = col_for(header, "type", "fuel", "technology")
    count, mw_sum = 0, 0.0
    types: dict[str, float] = {}
    for row in ws.iter_rows(min_row=hr + 1, values_only=True):
        if not any(c not in (None, "") for c in row):
            continue
        count += 1
        if mw_col is not None and mw_col < len(row):
            try:
                mw_sum += float(row[mw_col])
            except (TypeError, ValueError):
                pass
        if type_col is not None and type_col < len(row) and row[type_col]:
            t = str(row[type_col]).strip()[:24]
            try:
                types[t] = types.get(t, 0) + (
                    float(row[mw_col]) if mw_col is not None and row[mw_col] else 0
                )
            except (TypeError, ValueError):
                types[t] = types.get(t, 0)
    return {
        "rows": count,
        "mw_sum": round(mw_sum),
        "mw_col_header": header.get(mw_col or -1, "?"),
        "top_types_mw": dict(sorted(types.items(), key=lambda kv: -kv[1])[:8]),
    }


def main() -> None:
    raw = urllib.request.urlopen(
        urllib.request.Request(URL, headers={"User-Agent": UA}), timeout=60
    ).read()
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    sheets = {}
    for sn in wb.sheetnames:
        sheets[sn] = summarize_sheet(wb[sn])
    active = next(
        (v for k, v in sheets.items() if "queue" in k.lower() and "withdraw" not in k.lower()), {}
    )
    withdrawn = next((v for k, v in sheets.items() if "withdraw" in k.lower()), {})
    completed = next((v for k, v in sheets.items() if "complet" in k.lower()), {})
    ratio = (
        round(withdrawn.get("mw_sum", 0) / completed["mw_sum"], 2)
        if completed.get("mw_sum")
        else None
    )
    out = {
        "source": URL,
        "bytes": len(raw),
        "sheets": sheets,
        "active_queue_mw": active.get("mw_sum"),
        "completed_mw": completed.get("mw_sum"),
        "withdrawn_mw": withdrawn.get("mw_sum"),
        "withdrawn_rows": withdrawn.get("rows"),
        "active_rows": active.get("rows"),
        "withdrawn_to_completed_mw_ratio": ratio,
    }
    OUT_JSON.write_text(json.dumps(out, indent=2))
    write_md(out)
    print(
        f"wrote {OUT_JSON.relative_to(ROOT)} | active {out['active_queue_mw']}MW / "
        f"completed {out['completed_mw']}MW / withdrawn {out['withdrawn_mw']}MW "
        f"({out['withdrawn_rows']} projects)"
    )


BANNER = (ROOT / "analysis" / "economy_wide_fragility_map.md").read_text().split("\n")[2]


def write_md(out: dict) -> None:
    L = []
    L.append("# CAISO interconnection queue — power-buildout overbuild signal (keyless ISO data)")
    L.append("")
    L.append(BANNER)
    L.append("")
    L.append(
        f"Ingests the **FERC/ISO-queue dimension** keyless (proving it is *not* a wall): the "
        f"CAISO public interconnection queue. **Active {out['active_queue_mw']:,} MW** "
        f"({out['active_rows']} projects) vs **completed {out['completed_mw']:,} MW** vs "
        f"**withdrawn {out['withdrawn_mw']:,} MW across {out['withdrawn_rows']:,} projects** — a "
        f"withdrawn:completed MW ratio of **{out['withdrawn_to_completed_mw_ratio']}×**. The "
        f"withdrawn pile is the realized *announced-but-cancelled* buildout — the power-layer "
        f"analog of fiber-1999 overbuild (case zero's physical reality). "
        f"Machine-readable: [iso_queue_caiso.json](iso_queue_caiso.json)."
    )
    L.append("")
    L.append("## Per-sheet")
    L.append("")
    L.append("| sheet | projects | Σ MW | MW column |")
    L.append("|---|---:|---:|---|")
    for sn, s in out["sheets"].items():
        L.append(f"| {sn} | {s['rows']} | {s['mw_sum']:,} | {s['mw_col_header']} |")
    L.append("")
    L.append(
        "*One ISO (CAISO) as proof-of-reach; the other 6 ISOs (PJM/MISO/ERCOT/SPP/ISO-NE/NYISO) "
        "publish comparable keyless queues — laborious but not walled. The signal: a large "
        "withdrawn pipeline = capital/permitting committed to generation that never connected, "
        "the physical tell of demand-ahead-of-delivery in the AI-power buildout.*"
    )
    OUT_MD.write_text("\n".join(L) + "\n")


if __name__ == "__main__":
    main()
