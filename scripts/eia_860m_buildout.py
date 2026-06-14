"""EIA-860M generator inventory ingestion (off-EDGAR, KEYLESS — 5th source).

Found the real file by discovering the filename from the EIA index (my earlier 'needs
filename hunt = blocker' was wrong). Every US electric generator with operational STATUS:
operating / planned / under-construction / canceled / retired — a national buildout-vs-
attrition signal complementing the CAISO queue (case zero's physical-power layer).

Output: analysis/eia_860m_buildout.{json,md}
"""

from __future__ import annotations

import io
import json
import re
import urllib.request
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
OUT_JSON = ROOT / "analysis" / "eia_860m_buildout.json"
OUT_MD = ROOT / "analysis" / "eia_860m_buildout.md"
INDEX = "https://www.eia.gov/electricity/data/eia860m/"
BASE = "https://www.eia.gov"
UA = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/120 Safari/537.36"


def get(url: str) -> bytes:
    return urllib.request.urlopen(
        urllib.request.Request(url, headers={"User-Agent": UA}), timeout=90
    ).read()


def find_header(ws):
    best_row, best, best_map = 1, 0, {}
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=4, values_only=True), 1):
        sc = {
            i: str(c).strip()
            for i, c in enumerate(row)
            if c not in (None, "") and any(ch.isalpha() for ch in str(c))
        }
        if len(sc) > best:
            best, best_row, best_map = len(sc), ri, sc
    return best_row, best_map


def col(header, *keys):
    for i, name in header.items():
        if any(k in name.lower() for k in keys):
            return i
    return None


def main() -> None:
    idx = get(INDEX).decode("utf-8", "ignore")
    cands = re.findall(r'href="(/electricity/data/eia860m/(?:archive/)?xls/[a-z]+_generator\d{4}\.xlsx)"', idx, re.I)
    # index lists future placeholders that 404 -> add direct-constructed recent months as fallback
    months = ["april", "may", "march", "june", "february", "july", "january"]
    for sub in ("xls", "archive/xls"):
        for mo in months:
            for yr in ("2026", "2025"):
                cands.append(f"/electricity/data/eia860m/{sub}/{mo}_generator{yr}.xlsx")
    raw, file_url = None, None
    for path in dict.fromkeys(cands):
        try:
            data = get(BASE + path)
            if data[:2] == b"PK":  # genuine xlsx (zip), not an HTML 404
                raw, file_url = data, BASE + path
                break
        except Exception:
            continue
    if not raw:
        print(f"no valid EIA-860M xlsx among {len(cands)} candidates")
        return
    print(f"EIA-860M file: {file_url}")
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    sheets = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        hr, header = find_header(ws)
        mw_c = col(header, "nameplate", "capacity (mw)", "net summer")
        cnt, mw = 0, 0.0
        for row in ws.iter_rows(min_row=hr + 1, values_only=True):
            if not any(c not in (None, "") for c in row):
                continue
            cnt += 1
            if mw_c is not None and mw_c < len(row):
                try:
                    mw += float(row[mw_c])
                except (TypeError, ValueError):
                    pass
        sheets[sn] = {"rows": cnt, "mw_sum": round(mw), "mw_col": header.get(mw_c or -1, "?")}

    def pick(*keys):
        for sn, v in sheets.items():
            if any(k in sn.lower() for k in keys):
                return sn, v
        return None, {}

    op_n, op = pick("operating")
    pl_n, pl = pick("planned", "proposed")
    ca_n, ca = pick("cancel")
    re_n, rt = pick("retired")
    out = {
        "file": file_url,
        "sheets": sheets,
        "operating_mw": op.get("mw_sum"),
        "planned_mw": pl.get("mw_sum"),
        "canceled_mw": ca.get("mw_sum"),
        "retired_mw": rt.get("mw_sum"),
        "planned_to_operating_pct": (
            round(100 * pl.get("mw_sum", 0) / op["mw_sum"], 1) if op.get("mw_sum") else None
        ),
    }
    OUT_JSON.write_text(json.dumps(out, indent=2))
    write_md(out, {"operating": op_n, "planned": pl_n, "canceled": ca_n, "retired": re_n})
    print(
        f"wrote {OUT_JSON.relative_to(ROOT)} | operating {out['operating_mw']}MW / "
        f"planned {out['planned_mw']}MW / canceled {out['canceled_mw']}MW"
    )


BANNER = (ROOT / "analysis" / "economy_wide_fragility_map.md").read_text().split("\n")[2]


def write_md(out, names) -> None:
    L = []
    L.append("# EIA-860M generator buildout — national operating vs planned vs canceled (keyless)")
    L.append("")
    L.append(BANNER)
    L.append("")
    L.append(
        f"Ingests EIA-860M (every US generator + status), keyless. **Operating "
        f"{out['operating_mw']:,} MW** vs **planned {out['planned_mw']:,} MW** "
        f"(= {out['planned_to_operating_pct']}% of the operating fleet in the pipeline) vs "
        f"**canceled {out['canceled_mw']:,} MW** vs **retired {out['retired_mw']:,} MW** — the "
        f"national buildout-vs-attrition picture for the power layer feeding AI demand. "
        f"Machine-readable: [eia_860m_buildout.json](eia_860m_buildout.json)."
    )
    L.append("")
    L.append("| sheet | generators | Σ MW | MW column |")
    L.append("|---|---:|---:|---|")
    for sn, s in out["sheets"].items():
        L.append(f"| {sn} | {s['rows']} | {s['mw_sum']:,} | {s['mw_col']} |")
    L.append("")
    L.append(
        "*Complements CAISO's ~11x withdrawn:completed queue ratio: EIA-860M is the *built/"
        "building* reality (operating + firmly-planned), the queue is the *aspirational* funnel. "
        "Together they bound the AI-power buildout's demand-ahead-of-delivery gap with hard MW.*"
    )
    OUT_MD.write_text("\n".join(L) + "\n")


if __name__ == "__main__":
    main()
